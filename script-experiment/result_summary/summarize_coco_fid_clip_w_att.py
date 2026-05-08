#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Summarize COCO eval results (FID + CLIPScore) for w_att only into an Excel template (Sheet2).
- Reads run_dir/eval_coco/clip_summary.json and fid_summary.json
- Writes a clean table to Sheet2, keeps Sheet1 untouched
- Also exports a CSV for convenience

Expected run_dir naming (your convention):
  imgs/vis_sd14_GS_w_att_coco_seed12345
  imgs/vis_sd14_PRC_w_att_0_85_clip_coco_seed12345
  imgs/vis_sd14_T2S_w_att_coco_seed12345
  imgs/vis_sd14_TR_w_att_0_88_coco_seed12345
  ... similarly for sd15, sd21
"""

import os
import json
import argparse
from pathlib import Path
from typing import Dict, Any, List, Tuple

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


# -----------------------------
# Helpers
# -----------------------------
def load_json(p: Path) -> Dict[str, Any]:
    with p.open("r", encoding="utf-8") as f:
        return json.load(f)

def safe_get(d: Dict[str, Any], key: str, default=None):
    return d.get(key, default) if isinstance(d, dict) else default

def parse_run_dir_name(name: str) -> Dict[str, str]:
    """
    Parse:
      vis_sd14_GS_w_att_coco_seed12345
      vis_sd14_PRC_w_att_0_85_clip_coco_seed12345
      vis_sd14_TR_w_att_0_88_coco_seed12345
    Return fields:
      tag: sd14/sd15/sd21
      method: GS/PRC/T2S/TR
      variant: w_att
      seed: 12345
    """
    out = {"tag": "", "method": "", "variant": "", "seed": ""}

    if not name.startswith("vis_"):
        return out
    parts = name.split("_")

    # vis, sd14, METHOD..., ...
    # tag is parts[1]
    if len(parts) < 4:
        return out

    out["tag"] = parts[1]  # sd14/sd15/sd21

    # find variant
    # find variant (robust: don't rely on split tokens)
    if "_w_att_" in name:
        out["variant"] = "w_att"
    elif "_w_" in name:
        out["variant"] = "w"


    # seed token like seed12345
    for p in parts:
        if p.startswith("seed") and p[4:].isdigit():
            out["seed"] = p[4:]
            break

    # method inference: look for known keywords
    # PRC may appear as "PRC" then later "w" "att" ...
    if "PRC" in parts:
        out["method"] = "PRC"
    elif "GS" in parts:
        out["method"] = "GS"
    elif "T2S" in parts:
        out["method"] = "T2S"
    elif "TR" in parts:
        out["method"] = "TR"

    return out

def tag_to_model_name(tag: str) -> str:
    return {"sd14": "SD1.4", "sd15": "SD1.5", "sd21": "SD2.1"}.get(tag, tag)

def clear_sheet(ws):
    # remove all rows (safe way)
    if ws.max_row > 1:
        ws.delete_rows(1, ws.max_row)

def style_header_row(ws, row: int, ncols: int):
    bold = Font(bold=True)
    fill = PatternFill("solid", fgColor="DDDDDD")
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = bold
        cell.fill = fill
        cell.alignment = align

def autoset_col_width(ws, min_w: int = 10, max_w: int = 55):
    # simple heuristic by max string length
    for col in range(1, ws.max_column + 1):
        max_len = 0
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            s = str(v)
            max_len = max(max_len, len(s))
        w = max(min_w, min(max_w, max_len + 2))
        ws.column_dimensions[get_column_letter(col)].width = w


# -----------------------------
# Main
# -----------------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--template_xlsx", type=str, required=True,
                    help="Your existing Excel template (will keep Sheet1 unchanged)")
    ap.add_argument("--img_root", type=str, required=True,
                    help=".../experiment-1_19/imgs")
    ap.add_argument("--seed", type=str, default="12345")
    ap.add_argument("--variant", type=str, default="w_att", choices=["w_att", "w"])
    ap.add_argument("--out_xlsx", type=str, required=True,
                    help="Output summarized Excel path")
    ap.add_argument("--out_csv", type=str, required=True,
                    help="Output summarized CSV path")
    args = ap.parse_args()

    img_root = Path(args.img_root)
    seed = str(args.seed)
    variant = args.variant

    # Enumerate the 12 expected run dirs for w_att (or w)
    # NOTE: For TR w_att you use TR_w_att_0_88; for PRC w_att you use PRC_w_att_0_85_clip
    # We'll just glob by patterns to avoid hardcoding too tightly.


    run_dirs = sorted(img_root.glob(f"vis_*_{variant}*_coco_seed{seed}"))


    # Filter to only known methods
    rows = []
    missing = []

    for rd in run_dirs:
        info = parse_run_dir_name(rd.name)
        if info["variant"] != variant:
            continue
        if info["seed"] != seed:
            continue
        if info["method"] not in {"TR", "GS", "PRC", "T2S"}:
            continue

        clip_p = rd / "eval_coco" / "clip_summary.json"
        fid_p  = rd / "eval_coco" / "fid_summary.json"

        if (not clip_p.exists()) or (not fid_p.exists()):
            missing.append(str(rd))
            rows.append({
                "run_dir": str(rd),
                "model": tag_to_model_name(info["tag"]),
                "method": info["method"],
                "variant": variant,
                "clip_mean": None,
                "clip_std": None,
                "fid_point": None,
                "fid_bootstrap_std": None,
                "fid_ci95_low": None,
                "fid_ci95_high": None,
                "n": None,
            })
            continue

        clip = load_json(clip_p)
        fid = load_json(fid_p)

        bs = safe_get(fid, "bootstrap", {}) or {}
        rows.append({
            "run_dir": str(rd),
            "model": tag_to_model_name(info["tag"]),
            "method": info["method"],
            "variant": variant,
            "clip_mean": safe_get(clip, "mean"),
            "clip_std": safe_get(clip, "std"),
            "fid_point": safe_get(fid, "fid_point"),
            "fid_bootstrap_std": safe_get(bs, "std"),
            "fid_ci95_low": safe_get(bs, "ci95_low"),
            "fid_ci95_high": safe_get(bs, "ci95_high"),
            "n": safe_get(fid, "n"),
        })

    df = pd.DataFrame(rows)
    if len(rows) == 0:
        print("[FATAL] No run_dirs matched. Debug info:")
        print("  img_root =", img_root)
        print("  seed     =", seed)
        print("  variant  =", variant)
        # 打印 img_root 下可能的候选目录
        cand = sorted([p.name for p in img_root.glob(f"vis_*_{variant}*_coco_seed{seed}")])
        print("  candidates:", cand[:50], " ... total=", len(cand))
        raise SystemExit(2)

    # Order rows: SD1.4(TR,GS,PRC,T2S), SD1.5(...), SD2.1(...)
    model_order = {"SD1.4": 0, "SD1.5": 1, "SD2.1": 2}
    method_order = {"TR": 0, "GS": 1, "PRC": 2, "T2S": 3}
    df["_mo"] = df["model"].map(model_order).fillna(99).astype(int)
    df["_wo"] = df["method"].map(method_order).fillna(99).astype(int)
    df = df.sort_values(["_mo", "_wo"]).drop(columns=["_mo", "_wo"])

    # Save CSV
    out_csv = Path(args.out_csv)
    out_csv.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(out_csv, index=False)

    # Write into Excel template (Sheet2)
    tpl = Path(args.template_xlsx)
    wb = openpyxl.load_workbook(tpl)
    if "Sheet2" not in wb.sheetnames:
        wb.create_sheet("Sheet2")  # recommended creation method
    ws = wb["Sheet2"]  # keep Sheet1 unchanged

    clear_sheet(ws)

    headers = [
        "Model", "Method", "Variant",
        "CLIPScore_mean", "CLIPScore_std",
        "FID_point",
        "FID_bootstrap_std",
        "FID_ci95_low", "FID_ci95_high",
        "N",
        "run_dir"
    ]
    ws.append(headers)
    style_header_row(ws, row=1, ncols=len(headers))

    for _, r in df.iterrows():
        ws.append([
            r.get("model"),
            r.get("method"),
            r.get("variant"),
            r.get("clip_mean"),
            r.get("clip_std"),
            r.get("fid_point"),
            r.get("fid_bootstrap_std"),
            r.get("fid_ci95_low"),
            r.get("fid_ci95_high"),
            r.get("n"),
            r.get("run_dir"),
        ])

    # Freeze first row (set before heavy writing in write-only mode; here normal mode is fine)
    ws.freeze_panes = "A2"

    autoset_col_width(ws)

    out_xlsx = Path(args.out_xlsx)
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_xlsx)

    print("✅ Summarization done.")
    print(f"Saved Excel: {out_xlsx}")
    print(f"Saved CSV  : {out_csv}")
    if missing:
        print("\n[WARN] Missing eval files for some run_dirs (left as NA):")
        for m in missing:
            print("  -", m)


if __name__ == "__main__":
    main()
