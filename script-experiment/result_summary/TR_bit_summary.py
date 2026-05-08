#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import re
import glob
from typing import Dict, Tuple

import openpyxl
from openpyxl import Workbook


RUN_DIRS_DEFAULT = [
    "/home/yancy/work/dm_backdoor_latent_space/experiment-1_19/imgs/vis_sd14_TR_w_att_0_88_seed12345",
    "/home/yancy/work/dm_backdoor_latent_space/experiment-1_19/imgs/vis_sd14_TR_w_seed12345",
    "/home/yancy/work/dm_backdoor_latent_space/experiment-1_19/imgs/vis_sd15_TR_w_att_0_88_seed12345",
    "/home/yancy/work/dm_backdoor_latent_space/experiment-1_19/imgs/vis_sd15_TR_w_seed12345",
    "/home/yancy/work/dm_backdoor_latent_space/experiment-1_19/imgs/vis_sd21_TR_w_att_0_88_seed12345",
    "/home/yancy/work/dm_backdoor_latent_space/experiment-1_19/imgs/vis_sd21_TR_w_seed12345",
]

OUT_XLSX = "/home/yancy/work/dm_backdoor_latent_space/experiment-1_19/result_summary/TR_bit_accuracy.xlsx"


def find_detect_xlsx(run_dir: str) -> str:
    if not os.path.isdir(run_dir):
        raise FileNotFoundError(f"[run_dir] not found: {run_dir}")

    cands = sorted(glob.glob(os.path.join(run_dir, "*detect*.xlsx")))
    if len(cands) != 1:
        raise FileNotFoundError(
            f"[detect_xlsx] expect exactly 1 '*detect*.xlsx' in {run_dir}, got {len(cands)}: {cands}"
        )
    return cands[0]


def read_detect_rate(xlsx_path: str) -> float:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if "summary" not in wb.sheetnames:
        raise KeyError(f"[summary] sheet not found in {xlsx_path}. sheets={wb.sheetnames}")
    ws = wb["summary"]

    # Expect 2 columns: metric | value
    # Find row where metric == "detect_rate"
    detect_rate = None
    for (metric, value) in ws.iter_rows(min_row=1, values_only=True):
        if metric == "detect_rate":
            detect_rate = value
            break

    if detect_rate is None:
        raise KeyError(f"[detect_rate] not found in summary of {xlsx_path}")

    try:
        return float(detect_rate)
    except Exception as e:
        raise TypeError(f"[detect_rate] value not numeric in {xlsx_path}: {detect_rate} ({type(detect_rate)})") from e


def parse_tags(run_dir: str) -> Tuple[str, str]:
    """
    return (model_name, variant) where:
      model_name in {"SD1.4","SD1.5","SD2.1"}
      variant in {"w","w_att"}
    """
    name = os.path.basename(run_dir.rstrip("/"))

    m = re.search(r"vis_(sd14|sd15|sd21)_", name)
    if not m:
        raise ValueError(f"[parse] cannot parse model from dir name: {name}")
    sd = m.group(1)
    model = {"sd14": "SD1.4", "sd15": "SD1.5", "sd21": "SD2.1"}[sd]

    if "_w_att_" in name:
        variant = "w_att"
    elif "_w_" in name:
        variant = "w"
    else:
        raise ValueError(f"[parse] cannot parse variant (w/w_att) from dir name: {name}")

    return model, variant


def ensure_parent_dir(path: str) -> None:
    parent = os.path.dirname(path)
    if parent:
        os.makedirs(parent, exist_ok=True)


def write_summary_xlsx(out_xlsx: str, table: Dict[str, Dict[str, float]]) -> None:
    """
    table[model]["w"] = detect_rate
    table[model]["w_att"] = detect_rate
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "TR"

    header = [
        "Model",
        "w_detect_rate", "w_bit_accuracy",
        "w_att_detect_rate", "w_att_bit_accuracy",
    ]
    ws.append(header)

    order = ["SD1.4", "SD1.5", "SD2.1"]
    for model in order:
        if model not in table:
            raise KeyError(f"[missing] model {model} not collected")
        if "w" not in table[model] or "w_att" not in table[model]:
            raise KeyError(f"[missing] {model} missing w or w_att: {table[model]}")

        row_idx = ws.max_row + 1
        ws.append([
            model,
            table[model]["w"], "/",
            table[model]["w_att"], "/",
        ])

        # format detect_rate cells
        ws.cell(row=row_idx, column=2).number_format = "0.000"
        ws.cell(row=row_idx, column=4).number_format = "0.000"

    # basic column width
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18

    ensure_parent_dir(out_xlsx)
    wb.save(out_xlsx)


def main():
    run_dirs = RUN_DIRS_DEFAULT

    # Collect detect_rate
    table: Dict[str, Dict[str, float]] = {}
    seen = set()

    for rd in run_dirs:
        model, variant = parse_tags(rd)
        if (model, variant) in seen:
            raise RuntimeError(f"[dup] duplicated entry: {(model, variant)} from {rd}")
        seen.add((model, variant))

        xlsx = find_detect_xlsx(rd)
        dr = read_detect_rate(xlsx)

        table.setdefault(model, {})[variant] = dr
        print(f"[OK] {model:5s} {variant:5s} detect_rate={dr:.6f}  <- {xlsx}")

    # Strict: must cover all 3 models and both variants
    for model in ["SD1.4", "SD1.5", "SD2.1"]:
        if model not in table:
            raise RuntimeError(f"[missing] model {model} not found. got={list(table.keys())}")
        for variant in ["w", "w_att"]:
            if variant not in table[model]:
                raise RuntimeError(f"[missing] {model} missing {variant}. got={table[model].keys()}")

    write_summary_xlsx(OUT_XLSX, table)
    print(f"\n[DONE] saved: {OUT_XLSX}")


if __name__ == "__main__":
    main()
