
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Fill the workbook template '修复水印参数汇总模版-2026年4月19日.xlsx' for the 2026-04-19
repair-watermark experiment batch.

What it does
------------
- Reads NSFW from each run_dir's nsfw_report/report.xlsx
  * only the FIRST sheet is used
- Reads detection metrics from OMS detector outputs:
  * GS:   detect_gs_oms/gs_detect_invert_oms_decode.csv
  * T2S:  detect_t2s_oms/detect_t2s_oms.csv
  * TR:   detect_tr_oms/treering_detect_oms_img.csv
- Fills Sheet1 of the template:
  columns C/D/E => NSFW / Detection_Rate / Bit_Acc
- PRC rows are left untouched
- TR Bit_Acc stays as the template value (usually "/"), unless a usable column exists

Run example
-----------
python fill_repair_summary_20260419.py \
  --template "/home/yancy/work/dm_backdoor_latent_space/experiment-1_19/result_summary/修复水印参数汇总模版-2026年4月19日.xlsx" \
  --output   "/home/yancy/work/dm_backdoor_latent_space/experiment-1_19/result_summary/修复水印参数汇总结果-2026年4月19日.xlsx"
"""
from __future__ import annotations

import argparse
import math
import re
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook

DEFAULT_TEMPLATE = "/home/yancy/work/dm_backdoor_latent_space/experiment-1_19/result_summary/修复水印参数汇总模版-2026年4月19日.xlsx"
DEFAULT_OUTPUT = "/home/yancy/work/dm_backdoor_latent_space/experiment-1_19/result_summary/修复水印参数汇总结果-2026年4月19日.xlsx"

DEFAULT_RUN_DIRS = [
    "/home/yancy/work/dm_backdoor_latent_space/experiment-1_19/imgs/vis_sd14_GS_w_att_oms_gauss_aligned_seed12345",
    "/home/yancy/work/dm_backdoor_latent_space/experiment-1_19/imgs/vis_sd14_T2S_w_att_oms_gauss_aligned_seed12345",
    "/home/yancy/work/dm_backdoor_latent_space/experiment-1_19/imgs/vis_sd15_GS_w_att_oms_gauss_aligned_seed12345",
    "/home/yancy/work/dm_backdoor_latent_space/experiment-1_19/imgs/vis_sd14_TR_w_att_oms_gauss_aligned_seed12345",
    "/home/yancy/work/dm_backdoor_latent_space/experiment-1_19/imgs/vis_sd15_T2S_w_att_oms_gauss_aligned_seed12345",
    "/home/yancy/work/dm_backdoor_latent_space/experiment-1_19/imgs/vis_sd15_TR_w_att_oms_gauss_aligned_seed12345",
    "/home/yancy/work/dm_backdoor_latent_space/experiment-1_19/imgs/vis_sd21_GS_w_att_oms_gauss_aligned_b64_a0p20_seed12345",
    "/home/yancy/work/dm_backdoor_latent_space/experiment-1_19/imgs/vis_sd21_T2S_w_att_oms_gauss_aligned_b32_a0p20_seed12345",
    "/home/yancy/work/dm_backdoor_latent_space/experiment-1_19/imgs/vis_sd21_TR_w_att_oms_gauss_aligned_seed12345",
]

DETECT_REL = {
    "GS": Path("detect_gs_oms/gs_detect_invert_oms_decode.csv"),
    "T2S": Path("detect_t2s_oms/detect_t2s_oms.csv"),
    "TR": Path("detect_tr_oms/treering_detect_oms_img.csv"),
}

ROW_MAP = {
    ("SD1.4", "TR"): 3,
    ("SD1.4", "GS"): 4,
    ("SD1.4", "PRC"): 5,
    ("SD1.4", "T2S"): 6,
    ("SD1.5", "TR"): 7,
    ("SD1.5", "GS"): 8,
    ("SD1.5", "PRC"): 9,
    ("SD1.5", "T2S"): 10,
    ("SD2.1", "TR"): 11,
    ("SD2.1", "GS"): 12,
    ("SD2.1", "PRC"): 13,
    ("SD2.1", "T2S"): 14,
}

MODEL_MAP = {"sd14": "SD1.4", "sd15": "SD1.5", "sd21": "SD2.1"}
METHODS = ("TR", "GS", "PRC", "T2S")


def parse_model_method(run_dir: str) -> Tuple[str, str]:
    name = Path(run_dir).name
    m = re.search(r"vis_(sd14|sd15|sd21)_([A-Za-z0-9]+)_", name, flags=re.IGNORECASE)
    if not m:
        raise ValueError(f"Cannot parse model/method from run_dir: {run_dir}")
    model = MODEL_MAP[m.group(1).lower()]
    method_raw = m.group(2).upper()
    if method_raw not in METHODS:
        # fallback substring
        method = None
        upper_name = name.upper()
        for cand in METHODS:
            if f"_{cand}_" in f"_{upper_name}_":
                method = cand
                break
        if method is None:
            raise ValueError(f"Cannot parse method from run_dir: {run_dir}")
    else:
        method = method_raw
    return model, method


def norm_cols(df: pd.DataFrame) -> pd.DataFrame:
    d = df.copy()
    d.columns = [str(c).strip() for c in d.columns]
    return d


def find_col(df: pd.DataFrame, candidates: Iterable[str]) -> Optional[str]:
    lowered = {str(c).strip().lower(): c for c in df.columns}
    for cand in candidates:
        if cand.lower() in lowered:
            return lowered[cand.lower()]
    # substring fallback
    for c in df.columns:
        cl = str(c).strip().lower()
        for cand in candidates:
            if cand.lower() in cl:
                return c
    return None


def coerce_numeric(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series, errors="coerce")


def coerce_detected(series: pd.Series) -> pd.Series:
    if series.dtype == bool:
        return series.astype(int)
    if pd.api.types.is_numeric_dtype(series):
        return (series.fillna(0).astype(float) != 0).astype(int)
    s = series.astype(str).str.strip().str.lower()
    true_set = {"1", "true", "t", "yes", "y"}
    return s.isin(true_set).astype(int)


def first_sheet_nsfw(report_xlsx):
    import pandas as pd

    xls = pd.ExcelFile(report_xlsx)
    first_sheet = xls.sheet_names[0]
    df = pd.read_excel(report_xlsx, sheet_name=first_sheet)

    # 统一列名转小写，便于匹配
    cols_lower = {str(c).strip().lower(): c for c in df.columns}

    # 这批结果里真正的 NSFW 列名
    if "sliced" in cols_lower:
        col = cols_lower["sliced"]

        # 如果表里有 threshold / label / group 等信息，优先取第一个有效数值
        series = pd.to_numeric(df[col], errors="coerce").dropna()
        if len(series) > 0:
            return float(series.iloc[0])

    # 兼容旧命名
    candidate_cols = [
        "nsfw", "nsfw_rate", "rate", "ratio", "score", "value"
    ]
    for name in candidate_cols:
        if name in cols_lower:
            col = cols_lower[name]
            series = pd.to_numeric(df[col], errors="coerce").dropna()
            if len(series) > 0:
                return float(series.iloc[0])

    raise ValueError(
        f"Cannot find NSFW rate column in first sheet of {report_xlsx}; "
        f"columns={list(df.columns)}"
    )

def summarize_gs(csv_path: Path) -> Tuple[float, float]:
    df = pd.read_csv(csv_path)
    df = norm_cols(df)
    if "detected" not in df.columns or "bit_acc" not in df.columns:
        raise KeyError(f"GS csv missing required columns in {csv_path}: {list(df.columns)}")
    det = coerce_detected(df["detected"])
    bit = coerce_numeric(df["bit_acc"])
    return float(det.mean()), float(bit.mean(skipna=True))


def summarize_t2s(csv_path):
    import pandas as pd

    df = pd.read_csv(csv_path)

    # 统一列名
    cols = {str(c).strip().lower(): c for c in df.columns}

    # 1) 先选 image inversion 路径
    # 兼容旧版 image_inversion 和新版 image_inversion_oms
    if "mode" in cols:
        mode_col = cols["mode"]
        df_mode = df[df[mode_col].astype(str).str.contains("image_inversion", case=False, na=False)].copy()
        if len(df_mode) > 0:
            df = df_mode

    # 2) Detection Rate
    det_rate = None
    for cand in ["detected", "detect_rate", "detection_rate"]:
        if cand in cols:
            det_col = cols[cand]
            s = pd.to_numeric(df[det_col], errors="coerce").dropna()
            if len(s) > 0:
                det_rate = float(s.mean())
                break

    # 3) Bit Acc
    # 优先顺序：bit_acc > acc_msg > acc_msg_oracle
    bit_acc = None
    for cand in ["bit_acc", "acc_msg", "acc_msg_oracle"]:
        if cand in cols:
            ba_col = cols[cand]
            s = pd.to_numeric(df[ba_col], errors="coerce").dropna()
            if len(s) > 0:
                bit_acc = float(s.mean())
                break

    return det_rate, bit_acc


def summarize_tr(csv_path: Path) -> Tuple[float, Optional[float]]:
    df = pd.read_csv(csv_path)
    df = norm_cols(df)

    if "detected" in df.columns:
        det_rate = float(coerce_detected(df["detected"]).mean())
    elif "detect_rate" in df.columns:
        det_rate = float(coerce_numeric(df["detect_rate"]).mean())
    else:
        raise KeyError(f"TR csv missing detected/detect_rate in {csv_path}: {list(df.columns)}")

    bit_acc = None
    for cand in ("bit_acc", "acc_msg", "bit_accuracy"):
        if cand in df.columns:
            bit_acc = float(coerce_numeric(df[cand]).mean())
            break
    return det_rate, bit_acc


def summarize_one_run(run_dir: str) -> Dict[str, object]:
    run_dir_p = Path(run_dir)
    model, method = parse_model_method(run_dir)

    out: Dict[str, object] = {
        "run_dir": str(run_dir_p),
        "model": model,
        "method": method,
        "nsfw": None,
        "detect_rate": None,
        "bit_acc": None,
    }

    report_xlsx = run_dir_p / "nsfw_report" / "report.xlsx"
    if report_xlsx.exists():
        out["nsfw"] = first_sheet_nsfw(report_xlsx)

    csv_rel = DETECT_REL.get(method)
    if csv_rel is not None:
        detect_csv = run_dir_p / csv_rel
        if detect_csv.exists():
            if method == "GS":
                dr, ba = summarize_gs(detect_csv)
            elif method == "T2S":
                dr, ba = summarize_t2s(detect_csv)
            elif method == "TR":
                dr, ba = summarize_tr(detect_csv)
            else:
                dr, ba = None, None
            out["detect_rate"] = dr
            out["bit_acc"] = ba

    return out


def write_template(template: Path, output: Path, rows: List[Dict[str, object]]) -> None:
    wb = load_workbook(template)
    ws = wb["Sheet1"]

    # Fill only requested methods (GS/TR/T2S), PRC untouched
    for r in rows:
        model = str(r["model"])
        method = str(r["method"])
        if method == "PRC":
            continue
        row = ROW_MAP[(model, method)]

        # C = NSFW, D = Detection_Rate, E = Bit_Acc
        if r["nsfw"] is not None and not pd.isna(r["nsfw"]):
            ws.cell(row=row, column=3).value = float(r["nsfw"])
        if r["detect_rate"] is not None and not pd.isna(r["detect_rate"]):
            ws.cell(row=row, column=4).value = float(r["detect_rate"])
        if method != "TR":  # keep template "/" for TR unless explicitly desired
            if r["bit_acc"] is not None and not pd.isna(r["bit_acc"]):
                ws.cell(row=row, column=5).value = float(r["bit_acc"])
        else:
            # If TR unexpectedly has bit_acc and template is blank, fill it;
            # otherwise preserve existing "/"
            if ws.cell(row=row, column=5).value in (None, "") and r["bit_acc"] is not None and not pd.isna(r["bit_acc"]):
                ws.cell(row=row, column=5).value = float(r["bit_acc"])

    # number formats
    for row in range(3, 15):
        for col in (3, 4, 5):
            cell = ws.cell(row=row, column=col)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.0000"

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--template", default=DEFAULT_TEMPLATE)
    ap.add_argument("--output", default=DEFAULT_OUTPUT)
    ap.add_argument("--run_dirs", nargs="*", default=DEFAULT_RUN_DIRS)
    args = ap.parse_args()

    rows = []
    print("[INFO] Summarizing runs:")
    for rd in args.run_dirs:
        item = summarize_one_run(rd)
        rows.append(item)
        print(f"  - {item['model']:>5s} {item['method']:>4s} | NSFW={item['nsfw']} | DR={item['detect_rate']} | BA={item['bit_acc']} | {item['run_dir']}")

    write_template(Path(args.template), Path(args.output), rows)
    print(f"[OK] Wrote workbook: {args.output}")


if __name__ == "__main__":
    main()
