#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import re
import glob
from typing import Dict, Tuple

import openpyxl
from openpyxl import Workbook


# 6 个文件夹：sd14/sd15/sd21 × (delssc, delrepair)
RUN_DIRS_DEFAULT = [
    "/home/yancy/work/dm_backdoor_latent_space/experiment-1_19/imgs/vis_sd14_TR_w_att_0_88_delssc_seed12345",
    "/home/yancy/work/dm_backdoor_latent_space/experiment-1_19/imgs/vis_sd15_TR_w_att_0_88_delssc_seed12345",
    "/home/yancy/work/dm_backdoor_latent_space/experiment-1_19/imgs/vis_sd21_TR_w_att_0_88_delssc_seed12345",
    "/home/yancy/work/dm_backdoor_latent_space/experiment-1_19/imgs/vis_sd14_TR_w_att_0_88_delrepair_seed12345",
    "/home/yancy/work/dm_backdoor_latent_space/experiment-1_19/imgs/vis_sd15_TR_w_att_0_88_delrepair_seed12345",
    "/home/yancy/work/dm_backdoor_latent_space/experiment-1_19/imgs/vis_sd21_TR_w_att_0_88_delrepair_seed12345",
]

OUT_XLSX = "/home/yancy/work/dm_backdoor_latent_space/experiment-1_19/result_summary/TR_ablation_delssc_delrepair.xlsx"


def find_detect_xlsx(run_dir: str) -> str:
    if not os.path.isdir(run_dir):
        raise FileNotFoundError(f"[run_dir] not found: {run_dir}")

    cands = sorted(glob.glob(os.path.join(run_dir, "*detect*.xlsx")))
    if len(cands) != 1:
        raise FileNotFoundError(
            f"[detect_xlsx] expect exactly 1 '*detect*.xlsx' in {run_dir}, got {len(cands)}: {cands}"
        )
    return cands[0]


def read_detect_rate(xlsx_path: str) -> float:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if "summary" not in wb.sheetnames:
        raise KeyError(f"[summary] sheet not found in {xlsx_path}. sheets={wb.sheetnames}")
    ws = wb["summary"]

    detect_rate = None
    for (metric, value) in ws.iter_rows(min_row=1, values_only=True):
        if metric == "detect_rate":
            detect_rate = value
            break

    if detect_rate is None:
        raise KeyError(f"[detect_rate] not found in summary of {xlsx_path}")

    try:
        return float(detect_rate)
    except Exception as e:
        raise TypeError(f"[detect_rate] value not numeric in {xlsx_path}: {detect_rate} ({type(detect_rate)})") from e


def parse_tags(run_dir: str) -> Tuple[str, str]:
    """
    return (model_name, ablation) where:
      model_name in {"SD1.4","SD1.5","SD2.1"}
      ablation in {"delssc","delrepair"}
    """
    name = os.path.basename(run_dir.rstrip("/"))

    m = re.search(r"vis_(sd14|sd15|sd21)_", name)
    if not m:
        raise ValueError(f"[parse] cannot parse model from dir name: {name}")
    sd = m.group(1)
    model = {"sd14": "SD1.4", "sd15": "SD1.5", "sd21": "SD2.1"}[sd]

    if "_delssc_" in name:
        abla = "delssc"
    elif "_delrepair_" in name:
        abla = "delrepair"
    else:
        raise ValueError(f"[parse] cannot parse ablation (delssc/delrepair) from dir name: {name}")

    return model, abla


def ensure_parent_dir(path: str) -> None:
    parent = os.path.dirname(path)
    if parent:
        os.makedirs(parent, exist_ok=True)


def write_summary_xlsx(out_xlsx: str, table: Dict[str, Dict[str, float]]) -> None:
    """
    table[model]["delssc"] = detect_rate
    table[model]["delrepair"] = detect_rate
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "TR"

    header = [
        "Model",
        "delssc_detect_rate", "delssc_bit_accuracy",
        "delrepair_detect_rate", "delrepair_bit_accuracy",
    ]
    ws.append(header)

    order = ["SD1.4", "SD1.5", "SD2.1"]
    for model in order:
        if model not in table:
            raise KeyError(f"[missing] model {model} not collected")
        if "delssc" not in table[model] or "delrepair" not in table[model]:
            raise KeyError(f"[missing] {model} missing delssc or delrepair: {table[model]}")

        row_idx = ws.max_row + 1
        ws.append([
            model,
            table[model]["delssc"], "/",
            table[model]["delrepair"], "/",
        ])

        # format detect_rate cells
        ws.cell(row=row_idx, column=2).number_format = "0.000"
        ws.cell(row=row_idx, column=4).number_format = "0.000"

    # basic column width
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 22

    ensure_parent_dir(out_xlsx)
    wb.save(out_xlsx)


def main():
    run_dirs = RUN_DIRS_DEFAULT

    table: Dict[str, Dict[str, float]] = {}
    seen = set()

    for rd in run_dirs:
        model, abla = parse_tags(rd)
        if (model, abla) in seen:
            raise RuntimeError(f"[dup] duplicated entry: {(model, abla)} from {rd}")
        seen.add((model, abla))

        xlsx = find_detect_xlsx(rd)
        dr = read_detect_rate(xlsx)

        table.setdefault(model, {})[abla] = dr
        print(f"[OK] {model:5s} {abla:8s} detect_rate={dr:.6f}  <- {xlsx}")

    # Strict coverage: must cover all 3 models and both ablations
    for model in ["SD1.4", "SD1.5", "SD2.1"]:
        if model not in table:
            raise RuntimeError(f"[missing] model {model} not found. got={list(table.keys())}")
        for abla in ["delssc", "delrepair"]:
            if abla not in table[model]:
                raise RuntimeError(f"[missing] {model} missing {abla}. got={table[model].keys()}")

    write_summary_xlsx(OUT_XLSX, table)
    print(f"\n[DONE] saved: {OUT_XLSX}")


if __name__ == "__main__":
    main()
