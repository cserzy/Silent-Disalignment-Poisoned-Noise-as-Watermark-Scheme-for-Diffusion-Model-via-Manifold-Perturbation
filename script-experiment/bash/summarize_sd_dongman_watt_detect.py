#!/usr/bin/env python3
from __future__ import annotations

import csv
import json
import math
import re
import shutil
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


EXP_ROOT = Path("/home/yancy/work/dm_backdoor_latent_space/experiment-1_19")
TEMPLATE_PATH = EXP_ROOT / "result_summary/修复水印参数汇总模版-2026年4月19日.xlsx"
OUTPUT_XLSX = EXP_ROOT / "result_summary/SD-dongman-w_att水印检测汇总-2026年4月26日.xlsx"
OUTPUT_CSV = EXP_ROOT / "result_summary/SD-dongman-w_att水印检测汇总-2026年4月26日.csv"
OUTPUT_JSON = EXP_ROOT / "result_summary/SD-dongman-w_att水印检测汇总-2026年4月26日.json"
SCRIPT_PATH = Path(__file__).resolve()

RUN_DIRS = [
    EXP_ROOT / "imgs-number/vis_sd14_GS_w_att_dongman_seed12345",
    EXP_ROOT / "imgs-number/vis_sd14_T2S_w_att_dongman_seed12345",
    EXP_ROOT / "imgs-number/vis_sd15_GS_w_att_dongman_seed12345",
    EXP_ROOT / "imgs-number/vis_sd14_TR_w_att_0_88_dongman_seed12345",
    EXP_ROOT / "imgs-number/vis_sd15_T2S_w_att_dongman_seed12345",
    EXP_ROOT / "imgs-number/vis_sd15_TR_w_att_0_88_dongman_seed12345",
    EXP_ROOT / "imgs-number/vis_sd21_GS_w_att_dongman_seed12345",
    EXP_ROOT / "imgs-number/vis_sd21_T2S_w_att_dongman_seed12345",
    EXP_ROOT / "imgs-number/vis_sd21_TR_w_att_0_88_dongman_seed12345",
]

MODEL_PATTERNS = {
    "sd14": "SD1.4",
    "sd15": "SD1.5",
    "sd21": "SD2.1",
}
METHODS = ("GS", "T2S", "TR")
EXTENSIONS = (".json", ".csv", ".xlsx")


@dataclass
class CandidateResult:
    path: str
    detection_rate: float | None
    bit_acc: float | None
    score: int
    parser: str
    note: str


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = text.replace("%", "pct")
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def maybe_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        value = float(value)
        if math.isnan(value):
            return None
        return value
    text = str(value).strip()
    if not text:
        return None
    try:
        value = float(text)
    except ValueError:
        return None
    if math.isnan(value):
        return None
    return value


def mean(values: list[float]) -> float | None:
    cleaned = [v for v in values if v is not None]
    if not cleaned:
        return None
    return sum(cleaned) / len(cleaned)


def find_first_value(data: Any, keys: set[str]) -> Any:
    if isinstance(data, dict):
        for key, value in data.items():
            if normalize_text(key) in keys:
                return value
        for value in data.values():
            found = find_first_value(value, keys)
            if found is not None:
                return found
    elif isinstance(data, list):
        for item in data:
            found = find_first_value(item, keys)
            if found is not None:
                return found
    return None


def infer_model_method(run_dir: Path) -> tuple[str, str]:
    name = run_dir.name.lower()
    model = None
    for token, model_name in MODEL_PATTERNS.items():
        if token in name:
            model = model_name
            break
    method = None
    for candidate in METHODS:
        if f"_{candidate.lower()}_" in f"_{name}_":
            method = candidate
            break
    if not model or not method:
        raise ValueError(f"Failed to infer model/method from {run_dir}")
    return model, method


def list_candidate_files(run_dir: Path) -> list[Path]:
    files: list[Path] = []
    for path in run_dir.rglob("*"):
        if path.is_file() and path.suffix.lower() in EXTENSIONS:
            files.append(path)
    return sorted(files)


def read_csv_rows(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle))


def parse_gs_json(data: Any) -> tuple[float | None, float | None, str]:
    det = maybe_float(find_first_value(data, {"detected_rate", "detect_rate", "detected_mean"}))
    bit = maybe_float(find_first_value(data, {"bit_acc_mean", "bit_acc"}))
    if det is not None or bit is not None:
        return det, bit, "json-summary"
    return None, None, ""


def parse_t2s_json(data: Any) -> tuple[float | None, float | None, str]:
    det = maybe_float(find_first_value(data, {"detected_mean", "detected_rate"}))
    bit = maybe_float(find_first_value(data, {"acc_msg_oracle_mean"}))
    if det is not None or bit is not None:
        return det, bit, "json-summary"
    if isinstance(data, dict):
        detected_values: list[float] = []
        bit_values: list[float] = []
        for value in data.values():
            if not isinstance(value, dict):
                continue
            robustness = value.get("robustness")
            if not isinstance(robustness, dict):
                continue
            detected = maybe_float(robustness.get("detected"))
            acc_msg_oracle = maybe_float(robustness.get("acc_msg_oracle"))
            if detected is not None:
                detected_values.append(detected)
            if acc_msg_oracle is not None:
                bit_values.append(acc_msg_oracle)
        if detected_values or bit_values:
            return mean(detected_values), mean(bit_values), "json-detail-mean"
    return None, None, ""


def parse_tr_json(data: Any) -> tuple[float | None, float | None, str]:
    summary = data.get("summary", data) if isinstance(data, dict) else data
    det = maybe_float(find_first_value(summary, {"ok_rate", "detected_rate", "detect_rate"}))
    if det is not None:
        return det, None, "json-summary"
    if isinstance(data, dict):
        header = data.get("detail_header") or []
        rows = data.get("detail_rows") or []
        index = None
        normalized = [normalize_text(x) for x in header]
        if "ok" in normalized:
            index = normalized.index("ok")
        ok_values: list[float] = []
        for row in rows:
            if isinstance(row, dict):
                ok = maybe_float(row.get("ok"))
            elif index is not None and isinstance(row, list) and index < len(row):
                ok = maybe_float(row[index])
            else:
                ok = None
            if ok is not None:
                ok_values.append(ok)
        if ok_values:
            return mean(ok_values), None, "json-detail-mean"
    return None, None, ""


def parse_csv_for_method(path: Path, method: str) -> tuple[float | None, float | None, str]:
    rows = read_csv_rows(path)
    if not rows:
        return None, None, ""
    headers = {normalize_text(header): header for header in rows[0].keys()}
    detected_values: list[float] = []
    bit_values: list[float] = []

    if method == "GS":
        det_key = headers.get("detected")
        bit_key = headers.get("bit_acc") or headers.get("bit_acc_mean")
    elif method == "T2S":
        det_key = headers.get("detected")
        bit_key = headers.get("acc_msg_oracle") or headers.get("acc_msg_oracle_mean")
    else:
        det_key = headers.get("ok") or headers.get("detected")
        bit_key = None

    if det_key:
        detected_values = [maybe_float(row.get(det_key)) for row in rows]
        detected_values = [v for v in detected_values if v is not None]
    if bit_key:
        bit_values = [maybe_float(row.get(bit_key)) for row in rows]
        bit_values = [v for v in bit_values if v is not None]

    if detected_values or bit_values:
        return mean(detected_values), mean(bit_values), "csv-detail-mean"
    return None, None, ""


def parse_xlsx_for_method(path: Path, method: str) -> tuple[float | None, float | None, str]:
    wb = load_workbook(path, data_only=True)
    for ws in wb.worksheets:
        if ws.max_row >= 2 and ws.max_column >= 2:
            first_row = [normalize_text(ws.cell(1, c).value) for c in range(1, min(ws.max_column, 10) + 1)]
            if first_row[:2] == ["metric", "value"]:
                pairs = {}
                for r in range(2, ws.max_row + 1):
                    key = normalize_text(ws.cell(r, 1).value)
                    value = ws.cell(r, 2).value
                    if key:
                        pairs[key] = value
                if method == "TR":
                    det = maybe_float(pairs.get("ok_rate") or pairs.get("detected_rate") or pairs.get("detect_rate"))
                    if det is not None:
                        return det, None, f"xlsx-summary:{ws.title}"
                elif method == "GS":
                    det = maybe_float(pairs.get("detected_rate") or pairs.get("detect_rate") or pairs.get("detected_mean"))
                    bit = maybe_float(pairs.get("bit_acc_mean") or pairs.get("bit_acc"))
                    if det is not None or bit is not None:
                        return det, bit, f"xlsx-summary:{ws.title}"
                elif method == "T2S":
                    det = maybe_float(pairs.get("detected_mean") or pairs.get("detected_rate"))
                    bit = maybe_float(pairs.get("acc_msg_oracle_mean"))
                    if det is not None or bit is not None:
                        return det, bit, f"xlsx-summary:{ws.title}"

        headers = [normalize_text(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
        if method == "TR" and "ok" in headers:
            ok_col = headers.index("ok") + 1
            values = [maybe_float(ws.cell(r, ok_col).value) for r in range(2, ws.max_row + 1)]
            values = [v for v in values if v is not None]
            if values:
                return mean(values), None, f"xlsx-detail:{ws.title}"
        if method == "GS" and "detected" in headers and "bit_acc" in headers:
            det_col = headers.index("detected") + 1
            bit_col = headers.index("bit_acc") + 1
            det_values = [maybe_float(ws.cell(r, det_col).value) for r in range(2, ws.max_row + 1)]
            bit_values = [maybe_float(ws.cell(r, bit_col).value) for r in range(2, ws.max_row + 1)]
            det_values = [v for v in det_values if v is not None]
            bit_values = [v for v in bit_values if v is not None]
            if det_values or bit_values:
                return mean(det_values), mean(bit_values), f"xlsx-detail:{ws.title}"
        if method == "T2S" and "detected" in headers and "acc_msg_oracle" in headers:
            det_col = headers.index("detected") + 1
            bit_col = headers.index("acc_msg_oracle") + 1
            det_values = [maybe_float(ws.cell(r, det_col).value) for r in range(2, ws.max_row + 1)]
            bit_values = [maybe_float(ws.cell(r, bit_col).value) for r in range(2, ws.max_row + 1)]
            det_values = [v for v in det_values if v is not None]
            bit_values = [v for v in bit_values if v is not None]
            if det_values or bit_values:
                return mean(det_values), mean(bit_values), f"xlsx-detail:{ws.title}"
    return None, None, ""


def parse_candidate(path: Path, method: str) -> CandidateResult | None:
    lower_path = str(path).lower()
    detection_rate = None
    bit_acc = None
    parser = ""

    try:
        if path.suffix.lower() == ".json":
            with path.open("r", encoding="utf-8") as handle:
                data = json.load(handle)
            if method == "GS":
                detection_rate, bit_acc, parser = parse_gs_json(data)
            elif method == "T2S":
                detection_rate, bit_acc, parser = parse_t2s_json(data)
            else:
                detection_rate, bit_acc, parser = parse_tr_json(data)
        elif path.suffix.lower() == ".csv":
            detection_rate, bit_acc, parser = parse_csv_for_method(path, method)
        elif path.suffix.lower() == ".xlsx":
            detection_rate, bit_acc, parser = parse_xlsx_for_method(path, method)
    except Exception as exc:
        return CandidateResult(
            path=str(path),
            detection_rate=None,
            bit_acc=None,
            score=-100,
            parser="error",
            note=f"parse error: {exc}",
        )

    if detection_rate is None and bit_acc is None:
        return None

    score = 0
    filename = path.name.lower()
    if "summary" in filename:
        score += 40
    if "detect" in filename:
        score += 20
    if "report" in filename:
        score -= 30
    if "nsfw" in lower_path:
        score -= 60
    if "manifest" in filename:
        score -= 60
    if "treering" in filename:
        score += 10
    if path.suffix.lower() == ".xlsx":
        score += 12
    elif path.suffix.lower() == ".csv":
        score += 10
    elif path.suffix.lower() == ".json":
        score += 8
    if "summary" in parser:
        score += 30
    elif "detail" in parser:
        score += 10
    if method == "TR" and bit_acc is None:
        score += 5
    note = f"{parser}; score={score}"
    return CandidateResult(
        path=str(path),
        detection_rate=detection_rate,
        bit_acc=bit_acc,
        score=score,
        parser=parser,
        note=note,
    )


def choose_best_candidate(run_dir: Path, method: str) -> tuple[CandidateResult | None, list[CandidateResult], list[str]]:
    warnings: list[str] = []
    parsed: list[CandidateResult] = []
    for file_path in list_candidate_files(run_dir):
        candidate = parse_candidate(file_path, method)
        if candidate is not None:
            parsed.append(candidate)
    if not parsed:
        warnings.append("No parseable detection candidate found.")
        return None, [], warnings
    parsed.sort(key=lambda item: (item.score, item.detection_rate is not None, item.bit_acc is not None), reverse=True)
    best = parsed[0]
    if method == "TR" and best.bit_acc is not None:
        warnings.append("TR unexpectedly produced bit_acc; script will still write blank Bit_Acc.")
    if method == "TR":
        best.bit_acc = None
    return best, parsed, warnings


def locate_target_sheet(workbook) -> Any:
    for ws in workbook.worksheets:
        row2 = [normalize_text(ws.cell(2, c).value) for c in range(1, ws.max_column + 1)]
        if "model" in row2 and "detection_rate" in row2 and "bit_acc" in row2:
            return ws
    return workbook.active


def locate_group_columns(ws, group_name: str) -> dict[str, int]:
    target = normalize_text(group_name)
    merged_ranges = list(ws.merged_cells.ranges)
    for merged in merged_ranges:
        min_col, min_row, max_col, max_row = merged.bounds
        if min_row == 1 and normalize_text(ws.cell(min_row, min_col).value) == target:
            result: dict[str, int] = {}
            for col in range(min_col, max_col + 1):
                sub = normalize_text(ws.cell(2, col).value)
                if sub in {"nsfw", "detection_rate", "bit_acc"}:
                    result[sub] = col
            if "detection_rate" in result and "bit_acc" in result:
                return result

    current_group = None
    result = {}
    for col in range(1, ws.max_column + 1):
        value = normalize_text(ws.cell(1, col).value)
        if value:
            current_group = value
        if current_group == target:
            sub = normalize_text(ws.cell(2, col).value)
            if sub in {"nsfw", "detection_rate", "bit_acc"}:
                result[sub] = col
    if "detection_rate" in result and "bit_acc" in result:
        return result
    raise RuntimeError(f"Failed to locate columns for group={group_name}")


def locate_data_rows(ws) -> dict[tuple[str, str], int]:
    rows: dict[tuple[str, str], int] = {}
    current_model = None
    for row in range(3, ws.max_row + 1):
        model_cell = ws.cell(row, 1).value
        method_cell = ws.cell(row, 2).value
        if model_cell:
            current_model = str(model_cell).strip()
        method = str(method_cell).strip() if method_cell else None
        if current_model and method:
            rows[(current_model, method)] = row
    return rows


def preview_sheet(ws, max_rows: int = 8, max_cols: int = 12) -> list[list[Any]]:
    preview: list[list[Any]] = []
    for row in range(1, min(ws.max_row, max_rows) + 1):
        preview.append([ws.cell(row, col).value for col in range(1, min(ws.max_column, max_cols) + 1)])
    return preview


def write_outputs(records: list[dict[str, Any]], excel_info: dict[str, Any], warnings: list[str]) -> None:
    OUTPUT_CSV.parent.mkdir(parents=True, exist_ok=True)
    with OUTPUT_CSV.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=["model", "method", "run_dir", "detect_file", "detection_rate", "bit_acc", "note"],
        )
        writer.writeheader()
        for record in records:
            writer.writerow(
                {
                    "model": record["model"],
                    "method": record["method"],
                    "run_dir": record["run_dir"],
                    "detect_file": record["detect_file"],
                    "detection_rate": record["detection_rate"],
                    "bit_acc": record["bit_acc"],
                    "note": record["note"],
                }
            )

    payload = {
        "script_path": str(SCRIPT_PATH),
        "template_path": str(TEMPLATE_PATH),
        "output_excel": str(OUTPUT_XLSX),
        "output_csv": str(OUTPUT_CSV),
        "output_json": str(OUTPUT_JSON),
        "excel": excel_info,
        "records": records,
        "warnings": warnings,
    }
    OUTPUT_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def main() -> int:
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"Template not found: {TEMPLATE_PATH}")

    all_records: list[dict[str, Any]] = []
    global_warnings: list[str] = []

    for run_dir in RUN_DIRS:
        model, method = infer_model_method(run_dir)
        candidate_files = [str(path) for path in list_candidate_files(run_dir)]
        best, parsed_candidates, warnings = choose_best_candidate(run_dir, method)
        if warnings:
            global_warnings.extend([f"{run_dir.name}: {warning}" for warning in warnings])
        if best is None:
            record = {
                "model": model,
                "method": method,
                "run_dir": str(run_dir),
                "candidate_files": candidate_files,
                "parsed_candidates": [],
                "detect_file": None,
                "detection_rate": None,
                "bit_acc": None,
                "note": "No parseable detection result found.",
                "filled_excel": False,
            }
            all_records.append(record)
            print(f"[WARN] {model} / {method}: no parseable detection file")
            continue

        note = best.note
        if method == "TR":
            note = f"{note}; Bit_Acc left blank for TR"
        record = {
            "model": model,
            "method": method,
            "run_dir": str(run_dir),
            "candidate_files": candidate_files,
            "parsed_candidates": [asdict(item) for item in parsed_candidates],
            "detect_file": best.path,
            "detection_rate": best.detection_rate,
            "bit_acc": None if method == "TR" else best.bit_acc,
            "note": note,
            "filled_excel": False,
        }
        all_records.append(record)
        print(
            f"[FOUND] {model} / {method}: "
            f"detection_rate={record['detection_rate']} bit_acc={record['bit_acc']} "
            f"file={best.path}"
        )

    shutil.copy2(TEMPLATE_PATH, OUTPUT_XLSX)
    workbook = load_workbook(OUTPUT_XLSX)
    worksheet = locate_target_sheet(workbook)

    excel_info: dict[str, Any] = {
        "sheet": worksheet.title,
        "sheet_preview": preview_sheet(worksheet),
        "write_success": False,
        "columns": {},
        "row_map": {},
        "written_cells": [],
    }

    try:
        group_cols = locate_group_columns(worksheet, "w_att")
        row_map = locate_data_rows(worksheet)
        excel_info["columns"] = group_cols
        excel_info["row_map"] = {f"{k[0]}::{k[1]}": v for k, v in row_map.items()}

        for record in all_records:
            row = row_map.get((record["model"], record["method"]))
            if row is None:
                global_warnings.append(f"Excel row not found for {record['model']} / {record['method']}")
                continue
            det_col = group_cols["detection_rate"]
            bit_col = group_cols["bit_acc"]

            if record["detection_rate"] is not None:
                worksheet.cell(row, det_col).value = float(record["detection_rate"])
                worksheet.cell(row, det_col).number_format = "0.00%"
                excel_info["written_cells"].append(
                    {
                        "cell": worksheet.cell(row, det_col).coordinate,
                        "model": record["model"],
                        "method": record["method"],
                        "field": "Detection_Rate",
                        "value": record["detection_rate"],
                    }
                )
            if record["method"] != "TR" and record["bit_acc"] is not None:
                worksheet.cell(row, bit_col).value = float(record["bit_acc"])
                worksheet.cell(row, bit_col).number_format = "0.00%"
                excel_info["written_cells"].append(
                    {
                        "cell": worksheet.cell(row, bit_col).coordinate,
                        "model": record["model"],
                        "method": record["method"],
                        "field": "Bit_Acc",
                        "value": record["bit_acc"],
                    }
                )
            elif record["method"] == "TR":
                worksheet.cell(row, bit_col).value = None
                worksheet.cell(row, bit_col).number_format = "0.00%"
            record["filled_excel"] = record["detection_rate"] is not None
            print(
                f"[EXCEL] {record['model']} / {record['method']} -> "
                f"row={row} det_col={det_col} bit_col={bit_col}"
            )

        workbook.save(OUTPUT_XLSX)
        excel_info["write_success"] = True
    except Exception as exc:
        global_warnings.append(f"Excel write failed: {exc}")
        excel_info["write_success"] = False

    write_outputs(all_records, excel_info, global_warnings)

    print(f"[DONE] Excel: {OUTPUT_XLSX}")
    print(f"[DONE] CSV:   {OUTPUT_CSV}")
    print(f"[DONE] JSON:  {OUTPUT_JSON}")
    if global_warnings:
        print("[WARNINGS]")
        for warning in global_warnings:
            print(f"  - {warning}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
