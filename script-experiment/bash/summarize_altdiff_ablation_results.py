#!/usr/bin/env python3
import csv
import json
import math
import shutil
from pathlib import Path
from statistics import mean

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter


ROOT = Path("/home/yancy/work/dm_backdoor_latent_space")
EXP_ROOT = ROOT / "experiment-1_19"
IMGS_DIR = EXP_ROOT / "imgs"
SUMMARY_DIR = EXP_ROOT / "result_summary"

TEMPLATE_XLSX = SUMMARY_DIR / "第三批实验汇总-消融实验-2026年1月24日.xlsx"
OUT_XLSX = SUMMARY_DIR / "AltDiff-消融实验汇总-2026年4月21日.xlsx"
OUT_CSV = SUMMARY_DIR / "AltDiff-消融实验汇总-2026年4月21日.csv"
OUT_JSON = SUMMARY_DIR / "AltDiff-消融实验汇总-2026年4月21日.json"

MODEL_NAMES = {"AltDiff", "AltDiffusion"}
METHODS = ["GS", "TR", "T2S", "PRC"]
ABLATION_MAP = {
    "delssc": "w/o SSP",
    "delrepair": "w/o ARR",
}
GROUP_ORDER = ["w/o SSP", "w/o ARR"]
METRICS = ["NSFW", "Detection_Rate", "Bit_Acc"]


def norm_text(value):
    return str(value).strip().lower() if value is not None else ""


def safe_float(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)
    text = str(value).strip()
    if not text or text.lower() in {"nan", "none", "null", "/"}:
        return None
    is_percent = "%" in text
    text = text.replace("%", "")
    try:
        parsed = float(text)
    except Exception:
        return None
    return parsed / 100.0 if is_percent else parsed


def mean_numeric(values):
    clean = [v for v in values if v is not None]
    return mean(clean) if clean else None


def read_json(path):
    with path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def read_csv_rows(path):
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def pick_first_value(data, keys):
    if not isinstance(data, dict):
        return None, None
    for key in keys:
        value = safe_float(data.get(key))
        if value is not None:
            return value, key
    return None, None


def summarize_csv_column(rows, candidate_columns):
    fieldnames = list(rows[0].keys()) if rows else []
    for column in candidate_columns:
        if column not in fieldnames:
            continue
        values = [safe_float(row.get(column)) for row in rows]
        value = mean_numeric(values)
        if value is not None:
            return value, column
    return None, None


def cell_text(ws, row, col):
    cell = ws.cell(row, col)
    if isinstance(cell, MergedCell):
        return ""
    return norm_text(cell.value)


def print_sheet_preview(ws):
    print("[INFO] first sheet preview: first 20 rows x first 15 cols (non-empty only)")
    for row in range(1, 21):
        parts = []
        for col in range(1, 16):
            cell = ws.cell(row, col)
            if isinstance(cell, MergedCell):
                continue
            value = cell.value
            if value not in (None, ""):
                parts.append(f"{row},{col}={value}")
        if parts:
            print(" | ".join(parts))


def find_alt_method_rows(ws):
    rows = {}
    current_model = None
    for row in range(1, ws.max_row + 1):
        model_value = ws.cell(row, 1).value
        method_value = ws.cell(row, 2).value
        if model_value not in (None, ""):
            current_model = str(model_value).strip()
        if current_model in MODEL_NAMES and method_value not in (None, ""):
            method_name = str(method_value).strip()
            if method_name in METHODS:
                rows[method_name] = row
    return rows


def find_group_metric_columns(ws):
    found = {}
    max_scan_rows = min(ws.max_row, 6)
    max_scan_cols = min(ws.max_column, 30)
    for row in range(1, max_scan_rows + 1):
        for col in range(1, max_scan_cols + 1):
            label = cell_text(ws, row, col)
            if label not in {norm_text(x) for x in GROUP_ORDER}:
                continue
            group_name = "w/o SSP" if label == "w/o ssp" else "w/o ARR"
            metric_map = {}
            for scan_row in range(row, min(ws.max_row, row + 2) + 1):
                labels = [cell_text(ws, scan_row, col + offset) for offset in range(3)]
                if labels == ["nsfw", "detection_rate", "bit_acc"]:
                    metric_map = {
                        "NSFW": col,
                        "Detection_Rate": col + 1,
                        "Bit_Acc": col + 2,
                    }
                    break
            if metric_map:
                found[group_name] = metric_map
    return found


def format_pct_cell(cell, value):
    cell.value = value
    cell.number_format = "0.00%"


def infer_ablation_source(run_dir):
    name = run_dir.name.lower()
    for token, label in ABLATION_MAP.items():
        if token in name:
            return token, label
    return "", ""


def find_black_file(safeon_dir):
    preferred = safeon_dir / "black_detect" / "black_ratio_alt_summary.json"
    if preferred.exists():
        return preferred, []
    candidates = sorted(safeon_dir.rglob("*black*summary*.json"))
    if candidates:
        return candidates[0], [f"black file not at preferred path, used {candidates[0]}"]
    return None, [f"missing black summary under {safeon_dir}"]


def extract_black_rate(black_file):
    warnings = []
    if black_file is None:
        return None, {"source_type": "missing"}, warnings
    try:
        data = read_json(black_file)
    except Exception as exc:
        warnings.append(f"failed to read black summary: {black_file} | {exc}")
        return None, {"source_type": "json_error"}, warnings
    value, key = pick_first_value(data, ["black_rate", "safety_trigger_rate"])
    if value is None:
        warnings.append(f"black_rate field missing in {black_file}")
    return value, {"source_type": "json", "field": key, "keys": sorted(list(data.keys()))}, warnings


def find_gs_detect_file(run_dir):
    preferred = run_dir / "detect_gs_alt" / "gs_detect_invert_decode_alt_summary.json"
    if preferred.exists():
        return preferred, "json", []
    csv_candidates = sorted(run_dir.rglob("*.csv"))
    csv_candidates = [p for p in csv_candidates if "gs" in p.name.lower() and "detect" in p.name.lower()]
    if csv_candidates:
        return csv_candidates[0], "csv", [f"GS summary json missing, fallback to csv {csv_candidates[0]}"]
    return None, None, [f"missing GS detect result under {run_dir}"]


def extract_gs_metrics(run_dir):
    detect_file, source_type, warnings = find_gs_detect_file(run_dir)
    detail = {"source_type": source_type}
    if detect_file is None:
        return None, None, None, detail, warnings
    if source_type == "json":
        try:
            data = read_json(detect_file)
        except Exception as exc:
            warnings.append(f"failed to read GS summary json: {detect_file} | {exc}")
            return detect_file, None, None, detail, warnings
        det_value, det_key = pick_first_value(data, ["detected_rate", "detect_rate", "detected_mean"])
        bit_value, bit_key = pick_first_value(data, ["bit_acc_mean", "bit_acc"])
        if det_value is None:
            warnings.append(f"GS detection rate field missing in {detect_file}")
        if bit_value is None:
            warnings.append(f"GS bit_acc field missing in {detect_file}")
        detail.update({"det_field": det_key, "bit_field": bit_key, "keys": sorted(list(data.keys()))})
        return detect_file, det_value, bit_value, detail, warnings

    rows = read_csv_rows(detect_file)
    det_value, det_key = summarize_csv_column(rows, ["detected"])
    bit_value, bit_key = summarize_csv_column(rows, ["bit_acc"])
    if det_value is None:
        warnings.append(f"GS detected column missing in {detect_file}")
    if bit_value is None:
        warnings.append(f"GS bit_acc column missing in {detect_file}")
    detail.update({"det_field": det_key, "bit_field": bit_key, "rows": len(rows)})
    return detect_file, det_value, bit_value, detail, warnings


def find_tr_detect_file(run_dir):
    preferred = run_dir / "detect_tr_alt" / "treering_detect_alt_img.json"
    if preferred.exists():
        return preferred, []
    candidates = sorted(run_dir.rglob("treering*.json"))
    if candidates:
        return candidates[0], [f"TR detect json not at preferred path, used {candidates[0]}"]
    return None, [f"missing TR detect json under {run_dir}"]


def extract_tr_metrics(run_dir):
    detect_file, warnings = find_tr_detect_file(run_dir)
    detail = {"source_type": "json"}
    if detect_file is None:
        return None, None, None, detail, warnings
    try:
        data = read_json(detect_file)
    except Exception as exc:
        warnings.append(f"failed to read TR json: {detect_file} | {exc}")
        return detect_file, None, None, detail, warnings

    summary = data.get("summary", data if isinstance(data, dict) else {})
    det_value, det_key = pick_first_value(summary, ["ok_rate", "detected_rate", "detect_rate"])
    if det_value is None:
        header = data.get("detail_header") or []
        rows = data.get("detail_rows") or []
        ok_values = []
        if "ok" in header:
            idx = header.index("ok")
            ok_values = [safe_float(row[idx]) for row in rows if isinstance(row, list) and len(row) > idx]
            ok_values = [v for v in ok_values if v is not None]
        if ok_values:
            det_value = mean(ok_values)
            det_key = "detail_rows.ok"
    if det_value is None:
        warnings.append(f"TR detection rate missing in {detect_file}")
    detail.update({
        "det_field": det_key,
        "summary_keys": sorted(list(summary.keys())) if isinstance(summary, dict) else [],
        "detail_header": data.get("detail_header"),
    })
    return detect_file, det_value, None, detail, warnings


def find_t2s_detect_file(run_dir):
    preferred = run_dir / "detect_t2s_alt" / "t2s_detect_alt_results.json"
    if preferred.exists():
        return preferred, "json", []
    candidates = sorted(run_dir.rglob("t2s_detect_alt_results.json"))
    if candidates:
        return candidates[0], "json", [f"T2S preferred json missing, used {candidates[0]}"]
    csv_candidates = sorted(run_dir.rglob("t2s_detect_alt_results.csv"))
    if csv_candidates:
        return csv_candidates[0], "csv", [f"T2S json missing, fallback to csv {csv_candidates[0]}"]
    return None, None, [f"missing T2S detect result under {run_dir}"]


def extract_t2s_metrics(run_dir):
    detect_file, source_type, warnings = find_t2s_detect_file(run_dir)
    detail = {"source_type": source_type}
    if detect_file is None:
        return None, None, None, detail, warnings

    if source_type == "json":
        try:
            data = read_json(detect_file)
        except Exception as exc:
            warnings.append(f"failed to read T2S json: {detect_file} | {exc}")
            return detect_file, None, None, detail, warnings
        summary = data.get("summary_image_inversion_alt")
        summary_name = "summary_image_inversion_alt"
        if not isinstance(summary, dict):
            summary = data
            summary_name = "root"
        det_value, det_key = pick_first_value(summary, ["detected_mean", "detected_rate", "det_rate"])
        bit_value, bit_key = pick_first_value(summary, ["acc_msg_oracle_mean", "bit_accuracy_oracle", "acc_msg_oracle_mean"])
        if bit_value is None:
            bit_value, bit_key = pick_first_value(data, ["acc_msg_oracle_mean"])
            if bit_value is not None:
                summary_name = "root_fallback"
        if det_value is None:
            warnings.append(f"T2S detection rate missing in {detect_file}")
        if bit_value is None:
            warnings.append(f"T2S bit_acc missing in {detect_file}")
        detail.update({
            "summary_name": summary_name,
            "det_field": det_key,
            "bit_field": bit_key,
            "summary_keys": sorted(list(summary.keys())) if isinstance(summary, dict) else [],
        })
        return detect_file, det_value, bit_value, detail, warnings

    rows = read_csv_rows(detect_file)
    det_value, det_key = summarize_csv_column(rows, ["detected"])
    bit_value, bit_key = summarize_csv_column(rows, ["acc_msg_oracle"])
    if det_value is None:
        warnings.append(f"T2S detected column missing in {detect_file}")
    if bit_value is None:
        warnings.append(f"T2S acc_msg_oracle column missing in {detect_file}")
    detail.update({"det_field": det_key, "bit_field": bit_key, "rows": len(rows)})
    return detect_file, det_value, bit_value, detail, warnings


def prc_candidate_score(path):
    parent_name = path.parent.name.lower()
    score = 0
    if parent_name == "detect_prc_alt":
        score -= 100
    if "summary" in path.name.lower():
        score -= 10
    score += len(str(path))
    return score


def find_prc_detect_file(run_dir):
    json_candidates = sorted(run_dir.rglob("*.json"))
    json_candidates = [p for p in json_candidates if "prc" in p.name.lower() and "summary" in p.name.lower()]
    if json_candidates:
        chosen = sorted(json_candidates, key=prc_candidate_score)[0]
        warnings = []
        if len(json_candidates) > 1:
            warnings.append(f"multiple PRC summary json files found, chose {chosen}")
        return chosen, "json", warnings, [str(p) for p in json_candidates]

    csv_candidates = sorted(run_dir.rglob("*.csv"))
    csv_candidates = [
        p for p in csv_candidates
        if "prc" in p.name.lower() and ("detect" in p.name.lower() or "result" in p.name.lower())
    ]
    if csv_candidates:
        chosen = sorted(csv_candidates, key=prc_candidate_score)[0]
        warnings = []
        if len(csv_candidates) > 1:
            warnings.append(f"multiple PRC csv files found, chose {chosen}")
        return chosen, "csv", warnings, [str(p) for p in csv_candidates]
    return None, None, [f"missing PRC detect result under {run_dir}"], []


def extract_prc_metrics(run_dir):
    detect_file, source_type, warnings, candidates = find_prc_detect_file(run_dir)
    detail = {"source_type": source_type, "candidates": candidates}
    if detect_file is None:
        return None, None, None, detail, warnings

    if source_type == "json":
        try:
            data = read_json(detect_file)
        except Exception as exc:
            warnings.append(f"failed to read PRC summary json: {detect_file} | {exc}")
            return detect_file, None, None, detail, warnings
        det_value, det_key = pick_first_value(data, ["detected_mean", "detected_rate"])
        bit_value, bit_key = pick_first_value(data, ["bit_acc_mean", "msg_acc_mean", "message_acc_mean"])
        if det_value is None:
            warnings.append(f"PRC detection rate missing in {detect_file}")
        if bit_value is None:
            warnings.append(f"PRC bit_acc missing in {detect_file}")
        detail.update({"det_field": det_key, "bit_field": bit_key, "keys": sorted(list(data.keys()))})
        return detect_file, det_value, bit_value, detail, warnings

    rows = read_csv_rows(detect_file)
    det_value, det_key = summarize_csv_column(rows, ["detected", "detect", "is_detected"])
    bit_value, bit_key = summarize_csv_column(
        rows,
        ["bit_acc", "message_acc", "acc", "acc_msg", "msg_acc", "msg_bit_acc", "detect_bit_acc"],
    )
    if det_value is None:
        warnings.append(f"PRC detection column missing in {detect_file}")
    if bit_value is None:
        warnings.append(f"PRC bit_acc-like column missing in {detect_file}")
    detail.update({"det_field": det_key, "bit_field": bit_key, "rows": len(rows)})
    return detect_file, det_value, bit_value, detail, warnings


def extract_method_metrics(method, safeoff_dir, safeon_dir):
    warnings = []
    ablation_source, ablation = infer_ablation_source(safeoff_dir)
    black_file, black_warns = find_black_file(safeon_dir)
    warnings.extend(black_warns)
    black_rate, black_detail, black_extract_warns = extract_black_rate(black_file)
    warnings.extend(black_extract_warns)

    if method == "GS":
        detect_file, detection_rate, bit_acc, detect_detail, detect_warns = extract_gs_metrics(safeoff_dir)
    elif method == "TR":
        detect_file, detection_rate, bit_acc, detect_detail, detect_warns = extract_tr_metrics(safeoff_dir)
    elif method == "T2S":
        detect_file, detection_rate, bit_acc, detect_detail, detect_warns = extract_t2s_metrics(safeoff_dir)
    elif method == "PRC":
        detect_file, detection_rate, bit_acc, detect_detail, detect_warns = extract_prc_metrics(safeoff_dir)
    else:
        detect_file, detection_rate, bit_acc, detect_detail, detect_warns = (None, None, None, {}, [f"unsupported method {method}"])
    warnings.extend(detect_warns)

    note_parts = []
    if black_detail.get("field"):
        note_parts.append(f"black_field={black_detail['field']}")
    if detect_detail.get("det_field"):
        note_parts.append(f"det_field={detect_detail['det_field']}")
    if detect_detail.get("bit_field"):
        note_parts.append(f"bit_field={detect_detail['bit_field']}")
    if method == "TR":
        note_parts.append("TR Bit_Acc intentionally left blank")
    if method == "PRC" and detect_detail.get("candidates") and len(detect_detail["candidates"]) > 1:
        note_parts.append("multiple PRC candidates found")

    return {
        "model": "AltDiff",
        "method": method,
        "ablation": ablation,
        "ablation_source": ablation_source,
        "safeoff_run_dir": str(safeoff_dir),
        "safeon_run_dir": str(safeon_dir),
        "black_file": str(black_file) if black_file else "",
        "detect_file": str(detect_file) if detect_file else "",
        "black_rate": black_rate,
        "detection_rate": detection_rate,
        "bit_acc": bit_acc,
        "note": "; ".join(note_parts),
        "warnings": warnings,
        "black_detail": black_detail,
        "detect_detail": detect_detail,
    }


def build_runs():
    runs = []
    for method in METHODS:
        lower = method.lower()
        for ablation_source, ablation in ABLATION_MAP.items():
            safeoff_dir = IMGS_DIR / f"vis_alt_ablate_{lower}_{ablation_source}_seed12345"
            safeon_dir = IMGS_DIR / f"vis_alt_ablate_safeon_{lower}_{ablation_source}_seed12345"
            runs.append({
                "method": method,
                "ablation": ablation,
                "ablation_source": ablation_source,
                "safeoff_dir": safeoff_dir,
                "safeon_dir": safeon_dir,
            })
    return runs


def write_csv(records):
    fieldnames = [
        "model",
        "method",
        "ablation",
        "ablation_source",
        "safeoff_run_dir",
        "safeon_run_dir",
        "black_file",
        "detect_file",
        "black_rate",
        "detection_rate",
        "bit_acc",
        "note",
    ]
    with OUT_CSV.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for record in records:
            writer.writerow({key: record.get(key) for key in fieldnames})


def write_excel(records, preview_rows, row_map, col_map):
    shutil.copy2(TEMPLATE_XLSX, OUT_XLSX)
    wb = load_workbook(OUT_XLSX)
    ws = wb[wb.sheetnames[0]]

    excel_writes = []
    for record in records:
        method = record["method"]
        ablation = record["ablation"]
        row = row_map.get(method)
        group_cols = col_map.get(ablation, {})
        write_info = {
            "method": method,
            "ablation": ablation,
            "sheet_index": 1,
            "sheet_name": ws.title,
            "row": row,
            "columns": {},
        }
        if row is None:
            record["warnings"].append(f"missing AltDiff row for method {method} in template")
            excel_writes.append(write_info)
            continue
        if not group_cols:
            record["warnings"].append(f"missing column group for {ablation} in template")
            excel_writes.append(write_info)
            continue

        values = {
            "NSFW": record["black_rate"],
            "Detection_Rate": record["detection_rate"],
            "Bit_Acc": record["bit_acc"],
        }
        for metric, value in values.items():
            col = group_cols.get(metric)
            if col is None:
                record["warnings"].append(f"missing column for {ablation}/{metric} in template")
                continue
            cell = ws.cell(row, col)
            coord = f"{get_column_letter(col)}{row}"
            write_info["columns"][metric] = {"column": col, "cell": coord, "value": value}
            if metric == "Bit_Acc" and value is None:
                cell.value = None
                cell.number_format = "0.00%"
            else:
                format_pct_cell(cell, value)
        excel_writes.append(write_info)

    wb.save(OUT_XLSX)
    return {
        "sheet_index": 1,
        "sheet_name": ws.title,
        "preview": preview_rows,
        "row_map": row_map,
        "column_map": col_map,
        "writes": excel_writes,
    }


def main():
    if not TEMPLATE_XLSX.exists():
        raise FileNotFoundError(f"template not found: {TEMPLATE_XLSX}")

    wb = load_workbook(TEMPLATE_XLSX)
    ws = wb[wb.sheetnames[0]]
    print_sheet_preview(ws)

    preview_rows = []
    for row in range(1, 21):
        row_items = []
        for col in range(1, 16):
            cell = ws.cell(row, col)
            if isinstance(cell, MergedCell):
                continue
            if cell.value not in (None, ""):
                row_items.append({"row": row, "col": col, "value": cell.value})
        if row_items:
            preview_rows.append(row_items)

    row_map = find_alt_method_rows(ws)
    col_map = find_group_metric_columns(ws)
    print(f"[INFO] AltDiff row map: {row_map}")
    print(f"[INFO] column map: {col_map}")

    records = []
    for run in build_runs():
        record = extract_method_metrics(run["method"], run["safeoff_dir"], run["safeon_dir"])
        records.append(record)
        print(
            f"[INFO] {record['model']} / {record['method']} / {record['ablation']}: "
            f"black_rate={record['black_rate']} detection_rate={record['detection_rate']} bit_acc={record['bit_acc']}"
        )
        if record["warnings"]:
            for warning in record["warnings"]:
                print(f"[WARN] {record['method']} / {record['ablation']}: {warning}")

    excel_info = write_excel(records, preview_rows, row_map, col_map)
    write_csv(records)

    json_payload = {
        "template_xlsx": str(TEMPLATE_XLSX),
        "output_xlsx": str(OUT_XLSX),
        "output_csv": str(OUT_CSV),
        "output_json": str(OUT_JSON),
        "excel": excel_info,
        "records": [
            {
                "model": record["model"],
                "method": record["method"],
                "ablation": record["ablation"],
                "ablation_source": record["ablation_source"],
                "safeoff_run_dir": record["safeoff_run_dir"],
                "safeon_run_dir": record["safeon_run_dir"],
                "black_file": record["black_file"],
                "detect_file": record["detect_file"],
                "black_rate": record["black_rate"],
                "detection_rate": record["detection_rate"],
                "bit_acc": record["bit_acc"],
                "note": record["note"],
                "black_detail": record["black_detail"],
                "detect_detail": record["detect_detail"],
                "warnings": record["warnings"],
            }
            for record in records
        ],
        "warnings": [warning for record in records for warning in record["warnings"]],
    }
    with OUT_JSON.open("w", encoding="utf-8") as handle:
        json.dump(json_payload, handle, ensure_ascii=False, indent=2)

    print(f"[OK] wrote Excel: {OUT_XLSX}")
    print(f"[OK] wrote CSV:   {OUT_CSV}")
    print(f"[OK] wrote JSON:  {OUT_JSON}")


if __name__ == "__main__":
    main()
